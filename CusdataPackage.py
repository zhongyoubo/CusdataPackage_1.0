from doctest import master
from select import select
from sys import stderr, stdout
from threading import Thread
import tkinter as tk
from tkinter import Button, ttk
from tkinter import messagebox
from turtle import width
import openpyxl
from SSHManager import SSHManager
import sched
import datetime
import time



class LoginFrame(tk.Frame):
    def __init__(self,master=None):
     #Frame 是父类，需主动调用父类的构造器
        super().__init__(master) #显示调用父类的构造方法
        print("LoginFrame init")
        self.pack(pady=10)
        self.master = master
        self.hostname = tk.StringVar()
        self.hostname.set(g_HostName)
        self.username = tk.StringVar()
        self.username.set(g_UserName)
        self.password = tk.StringVar()
        self.password.set(g_PassWord)

        self.createWidgets()

    def createWidgets(self):
        print("LoginFrame createWidgets")
        self.hostname_label = ttk.Label(self,text="HostName",width=10)
        self.hostname_label.grid(row=0,column=0,sticky="we")
        
        self.hostname_entry = ttk.Entry(self,textvariable=self.hostname)
        self.hostname_entry.grid(row=0,column=1,sticky="we")
        self.username_label = ttk.Label(self,text="UserName")
        self.username_label.grid(row=1,column=0,sticky="we",pady=5)
        
        self.username_entry = ttk.Entry(self,textvariable=self.username)
        self.username_entry.grid(row=1,column=1,sticky="we",pady=5)
        self.password_label = ttk.Label(self,text="Password")
        self.password_label.grid(row=2,column=0,sticky="we")
        
        self.password_entry = ttk.Entry(self,textvariable=self.password)
        self.password_entry.grid(row=2,column=1,sticky="we")
        self.login_button = ttk.Button(self,text="Login",command=self.login_check)
        self.login_button.grid(row=0,column=3,rowspan=3,sticky="ns",padx=20)

    def login_check(self):
        print("login check,hostname:",self.hostname.get(),"username:",self.username.get(),"password:",self.password.get())
        #global g_SSHManager
        res_code = g_SSHManager.ssh_connect(hostname=g_HostName,username=g_UserName,password=g_PassWord)
        global g_Login_Success
        if res_code == 1000:
            print("Login Success")
            g_Login_Success = True
            messagebox.showinfo(title="Message",message="Login Success")
        else:  
            print("Login Fail")
            g_Login_Success = False   
            messagebox.showerror(title="Error",message="Login Fail")  


class BuildFrame(tk.Frame): #继承Frame类
    def __init__(self,master=None):
        #Frame 是父类，需主动调用父类的构造器
        super().__init__(master) #显示调用父类的构造方法
        print("BuildFrame init")
        self.pack()
        self.master = master
        self.region = tk.StringVar()
        self.region.set(g_Region)
        self.version = tk.StringVar()
        self.version.set(g_Version)
        self.createWidgets()

    def createWidgets(self):
        print("BuildFrame createWidgets")
        self.region_label = ttk.Label(self,text="Region")
        self.region_label.grid(row=0,column=0,sticky="we")
        
        self.region_cbb = ttk.Combobox(self,textvariable=self.region,values=g_Region_List) #下拉列表的值
        self.region_cbb.bind("<<ComboboxSelected>>",self.select_region)
        self.region_cbb.grid(row=0,column=1,sticky="we")
        self.version_label = ttk.Label(self,text="Version")
        self.version_label.grid(row=1,column=0,sticky="we")
        
        self.version_cbb = ttk.Combobox(self,textvariable=self.version,values=list(g_Version_Path.keys()))
        self.version_cbb.bind("<<ComboboxSelected>>",self.select_version)
        self.version_cbb.grid(row=1,column=1,sticky="we")

        self.build_button = ttk.Button(self,text="Build",command=self.build_clicked)
        self.build_button.grid(row=0,column=3,padx=20)
        self.cancel_button = ttk.Button(self,text="Cancel",command=self.cancel_clicked)
        self.cancel_button.grid(row=1,column=3,padx=20)

    def select_region(self,event):
        print("Region",self.region.get(),"is selected")
        global g_Region
        g_Region = self.region.get()

    def select_version(self,event):
        print("Version",self.version.get(),"is Selected")
        global g_Version
        g_Version = self.version.get()


    def build_clicked(self):
        print("Start Build, Region:",g_Region," Version:",g_Version,"CodePath:",g_Version_Path[g_Version])
        if g_Login_Success: 
            if not self.isCodeBuiltAll():return
            stdout = g_SSHManager.ssh_exec_command(f"bash {g_Scripts_Path}/Start_Cusdata_Auto_Package.sh {g_Region} {g_Version_Path[g_Version]}")
            print("stdout:",stdout)
            g_Build_Checking = True
            self.build_button["text"] = "Building"
            self.build_button["state"] = "disabled"
            self.scheduler_check = g_Scheduler.enter(10,1,self.build_check,()) # 每隔10秒执行一次，优先级为1，需要执行的目标函数为build_check，这里build_check函数无需传参
            self.thread_check = Thread(target=g_Scheduler.run)
            self.thread_check.start()
        else:
            print("Please Login First")
            messagebox.showerror(title="Error",message="Please Login First")
            return   
             
    def cancel_clicked(self):
        print("Cancel Build")
        stdout = g_SSHManager.ssh_exec_command(f"bash {g_Scripts_Path}/Stop_Cusdata_Auto_Package.sh")
        print("Cancel Result:",stdout)
        self.build_button["text"] = "Build"
        self.build_button["state"] = "normal"
        g_Scheduler.cancel(self.build_check)

    
    def build_check(self):
        self.scheduler_check = g_Scheduler.enter(10,1,self.build_check,())
        print("cur time:",datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        result = g_SSHManager.ssh_exec_command('[ -f /home/SDD_DISK/publisherB/SA_Build_Cusdata_Success ] && echo yes || echo no')
        print('build_check result1:',result.strip("\r\n"))
        if result.strip("\r\n") == "yes":
            print("Cusdata Package Success")
            messagebox.showinfo(title="Message",message="Cusdata Package Success")
            self.build_button["text"] = "Build"
            self.build_button["state"] = "normal"
            g_Scheduler.cancel(self.scheduler_check)
            g_Build_Checking = False
            return
        result = g_SSHManager.ssh_exec_command('[ -f /home/SDD_DISK/publisherB/SA_Build_Cusdata_Error ] && echo yes || echo no')
        print('build_check result2:',result.strip("\r\n"))
        if result.strip("\r\n") == "yes":
            print("Cusdata Package Fail")
            messagebox.showinfo(title="Message",message="Cusdata Package Fail")
            self.build_button["text"] = "Build"
            self.build_button["state"] = "normal"
            g_Scheduler.cancel(self.scheduler_check)
            g_Build_Checking = False
            return

    def isCodeBuiltAll(self):   
        result = g_SSHManager.ssh_exec_command(f'[ -f {g_Version_Path[g_Version]}/release/out/mediatek_linux/output/upgrade_image.pkg ] && echo yes || echo no')
        print('isCodeBuiltAll result:',{result})
        if result.strip('\r\n') == 'no':
            print('the code has not been built all')
            messagebox.showerror('Error','the code has not been built all')
            return False
        return True
        
        

 #从DataConfig.xlsx获取信息
def loadDataConfig():
    wb = openpyxl.load_workbook("./DataConfig.xlsx")
    ws = wb.active
    global g_Region_List
    g_Region_List = ["AP","EU","SA","NA"]
    global g_HostName,g_UserName,g_PassWord,g_Project,g_Region,g_Version #软件或者tarball版本，如R0,R1
    g_HostName = ws["B1"].value
    g_UserName = ws["B2"].value
    g_PassWord = ws["B3"].value
    g_Project = ws["B4"].value
    g_Region = ws["B5"].value
    
    print("g_HostName:",g_HostName,"g_UserName:",g_UserName,"g_PassWord:",g_PassWord)
    print("g_Project:",g_Project,"g_Region:",g_Region)
    global g_Version_Path #软件版本与CodePath映射表
    g_Version_Path = {}
    for row in range(6,ws.max_row+1):
        print(ws[f"A{row}"].value,ws[f"B{row}"].value)
        v = ws[f"A{row}"].value
        p = ws[f"B{row}"].value
        g_Version_Path.update({v:p})
    print("g_Version_Path:",g_Version_Path)
    g_Version = list(g_Version_Path.keys())[0] #先默认在第一个软件版本路径





if __name__ == "__main__":
    print("lanuch application of 9216 Cusdata Package Tool")
    root = tk.Tk()
    screen_w,screen_h = root.maxsize()
    win_w = 400
    win_h = 180
    x,y = int((screen_w-win_w)/2),int((screen_h-win_h)/2)
    print("screen_w",screen_w,"screen_h",screen_h,"win_w",win_w,"win_h",win_h,"x",x,"y",y)
    root.geometry("{}x{}+{}+{}".format(win_w,win_h,x,y))
    root.title('9216 Cusdata Package Tool')
    loadDataConfig()
    global g_SSHManager,g_Login_Success,g_Scripts_Path,g_Scheduler,g_Build_Checking
    g_Login_Success = False   #用于判断是否登录
    g_Build_Checking = False  #用于判断是否在build
    g_Scripts_Path = "/home/SDB_DISK/publisherB/SDD_DISK/MT9216_TARBALL/scripts"
    g_SSHManager = SSHManager()
    g_Scheduler = sched.scheduler(time.time,time.sleep)
    login_frame = LoginFrame(master=root)
    build_frame = BuildFrame(master=root)

    root.mainloop()
